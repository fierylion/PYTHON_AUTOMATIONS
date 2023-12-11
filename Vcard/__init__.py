from typing import Any
from openpyxl import load_workbook
#defined fields for vcard and excel
#fields = ['full_name','first_name', 'last_name', 'email', 'phone', 'address', 'company', 'title']
#full_name means the code will splits the name ie name.split('') and obtain:
#first_name = full_name[0], last_name=full_name[-1]

class VcardExcel():
    
    #fields maps can be like {'full_name': 'Full Name', email:'Emails'}
    def __init__(self, excel_sheet, field_maps, vcf_file,slug=''):
        self.slug = slug
        self.fields= ['full_name','first_name', 'last_name', 'email', 'phone', 'address', 'company', 'title']
        self.excel_sheet = excel_sheet
        self.field_maps = field_maps
        self.vcf_file = vcf_file
    @staticmethod
    def format_phone_number(phone_number):
        if not phone_number:
            return phone_number
        phone_number = str(phone_number)
     
        if phone_number.startswith('0'):
            phone_number = f'+255{phone_number[1:]}'
        if phone_number.startswith('7'):
            phone_number = f'+255{phone_number}'
        if phone_number.startswith('255'):
            phone_number = f'+{phone_number}'
        if phone_number.startswith('+'):
            return phone_number
        
        return phone_number
    @staticmethod
    def capitalize(s):
        if(len(s)<1):
            return s
        return s[0].upper() + s[1:]
    def generate_vcard(self):
        wb = load_workbook(self.excel_sheet)
        ws = wb.active
        col_map ={}
        #create map of columns to fields map
        for ky in self.field_maps.keys():
            if ky not in self.fields:
                raise Exception(f'{ky} is not a valid field')
            for i in range(1, ws.max_column+1):
                if ws.cell(row=1, column=i).value == self.field_maps[ky]:
                    col_map[ky] = i
        for i in range(2, ws.max_row+1):
           field_values = {}
           for ky in col_map.keys():
               val = ws.cell(row=i, column=col_map[ky]).value
               field_values[ky] = val
           
           
           
           if(field_values.get('full_name')):
                full_name = field_values['full_name']
                
                first_name =self.capitalize( full_name.split(' ')[0])
                
                last_name = self.capitalize(full_name.split(' ')[-1])
                
                
                vcard = self.make_vcard(
                    first_name=first_name,
                    last_name=last_name +f'_{self.slug}{int(ws.cell(row=i, column=9).value)}', #temporary fix
                    email=field_values.get('email', ''),
                    phone=self.format_phone_number(
                        field_values.get('phone', '')),
                    address=field_values.get('address', ''),
                    title=field_values.get('title', ''),
                    company=field_values.get('company', '')
                )
                self.write_vcard(self.vcf_file, vcard)
                
           else:
               vcard = self.make_vcard(
                   first_name=self.capitalize(field_values.get('first_name', '')),
                   last_name=self.capitalize( field_values.get(
                       'last_name', '')) + f'_{self.slug}{int(ws.cell(row=i, column=9).value)}', #temporary fix

                   email=field_values.get('email', ''),
                   phone=self.format_phone_number(field_values.get('phone', '')),
                   address=field_values.get('address', ''),
                   title=field_values.get('title', ''),
                   company=field_values.get('company', '')
               )
               self.write_vcard(self.vcf_file, vcard)
            
            

            

    
        
    
    def write_vcard(self, f, vcard):
        with open(f, 'a') as f:
            f.writelines([l + '\n' for l in vcard])
    
    def make_vcard(
        self,
        first_name,
        last_name,
        company,
        title,
        phone,
        address,
        email):
        # address_formatted = ';'.join([p.strip() for p in address.split(',')])
        return [
            'BEGIN:VCARD',
            'VERSION:2.1',
            f'N:{last_name};{first_name}',
            f'FN:{first_name} {last_name}',
            f'ORG:{company}',
            f'TITLE:{title}',
            f'EMAIL;PREF;INTERNET:{email}',
            f'TEL;WORK;VOICE:{phone}',
            f'ADR;WORK;PREF:;;{address}',
            f'REV:1',
            'END:VCARD'
        ]


vcard = VcardExcel('upskilling.xlsx', {
                   'full_name': 'Full Name', 'email': 'Email', 'phone': 'Your WhatsApp Phone Number'}, 'upskilling_contact.vcf', 'UPS')
vcard.generate_vcard()