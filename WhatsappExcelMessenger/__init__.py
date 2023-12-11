from alright import WhatsApp
from openpyxl import load_workbook
class WhatsappMessenger():
    def __init__(self, excel_sheet, phone_col, name_col):
        self.name_col = name_col
        self.excel_sheet = excel_sheet
        self.phone_col = phone_col
        self.messenger = WhatsApp()
    @staticmethod
    def format_phone_number(phone_number):
        if not phone_number:
            print('no phone number')
        phone_number = str(phone_number)
        if phone_number.startswith('0'):
            phone_number = f'255{phone_number[1:]}'
        if phone_number.startswith('7'):
            phone_number = f'255{phone_number}'
        if phone_number.startswith('255'):
            phone_number = f'{phone_number}'
        if phone_number.startswith('+'):
            return phone_number[1:]
        
        return phone_number
    @staticmethod
    def capitalize(s):
        if(len(s)<1):
            return s
        return s[0].upper() + s[1:]
    @staticmethod
    def format_name(name):
        if not name:
            print('no name')
            return ''
        names = name.split(' ')
        res_names = []
        for name in names:
            res_names.append(WhatsappMessenger.capitalize(name))
        return ' '.join(res_names)



    def send_messages(self):
        wb = load_workbook(self.excel_sheet)
        ws = wb.active
        last_number = '255763881843'
        continue_sending = False
        obtained_number=False
        no_of_jumps=0
        for i in range(2, ws.max_row+1):
            phone_number = self.format_phone_number(ws.cell(row=i, column=self.phone_col).value)
            if(obtained_number and not continue_sending):
                no_of_jumps+=1
                if(no_of_jumps==1):
                    continue_sending=True
                    continue
            if (last_number == phone_number):
                obtained_number=True
                continue
            if not continue_sending:
                continue
            name = self.format_name(ws.cell(row=i, column=self.name_col).value)
            message = f'Hello {name}, Congratulations, You have been selected for the UDICTI upskilling program 2023/2024. \n Program Starts: 13th Wednesday,2023 \n Venue: UDICTI HUB, BLOCK B, (B109). \n If you haven\'t been added to the whatsapp group, please reply to this message.'
            if phone_number:
                try:
                    self.messenger.find_user(phone_number)
                except:
                    print('user not found')
                    continue
                self.messenger.send_message(message)
            else:
                print(name, phone_number, 'no phone number')
             
        
    